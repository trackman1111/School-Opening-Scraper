import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
from datetime import datetime


def main():
    modifiedDate = download_xslx()
    #print("OH - Downloaded CSV")
    copy_to_new_csv(modifiedDate)
    #print("OH - Wrote CSV")


def download_xslx():
    # Get html of page
    url = "http://education.ohio.gov/Topics/Reset-and-Restart"
    html = requests.get(url).content
    soup = BeautifulSoup(html, 'html.parser')

    # Get the most recent update link
    path = soup.select_one('div[id="main-content"]').find("a", string="this data compilation")['href']
    date = path[-10:-6]
    dataUrl = "http://education.ohio.gov" + str(path)
    # Retrieve cvs file
    originalFile = requests.get(dataUrl)
    open('temp/OhioOriginal.xlsx', 'wb').write(originalFile.content)
    return date


def copy_to_new_csv(modifiedDate):
    wb = load_workbook('temp/OhioOriginal.xlsx')
    districtSheet = wb['Model']
    inputRow = 0
    df = pd.DataFrame(
        columns=['district irn', 'district', 'county',
                 'current model', 'report date', 'date scraped'])

    for row in districtSheet.iter_rows(values_only=True):
        if inputRow == 0:  # Skip column headers
            inputRow += 1
            continue
        districtIRN = row[0]
        districtName = row[1]
        countyName = row[2]
        currentModel = row[3]
        if districtName is None:  # End of input file
            break

        newRow = pd.Series(data={'district irn': districtIRN, 'district': districtName,
                                         'county': countyName, 'current model': currentModel,
                                         'report date': modifiedDate, 'date scraped': date.today()})

        df = df.append(newRow, ignore_index=True)

        inputRow += 1  # End for
    df.to_csv('out/OH_' + datetime.now().strftime('%Y%m%d') + '.csv', index=False)  # Copy dataframe to CSV

#main()
