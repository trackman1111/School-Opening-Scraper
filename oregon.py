import requests
import urllib
import csv
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date


def main():
    download_xslx()
    copy_to_new_csv()

def download_xslx():
    # Get html of page
    url = "https://www.oregon.gov/ode/students-and-family/healthsafety/Pages/2020-21-School-Status.aspx"
    html = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(html, 'html.parser')

    # Get the most recent update link
    path = soup.body.main.div.contents[9].li.a['href']
    dataUrl = "https://www.oregon.gov/" + path

    # Retrieve cvs file
    urllib.request.urlretrieve(dataUrl, './OregonOriginal.xlsx')
    print("Downloaded OR xlsx")


def copy_to_new_csv():
    wb = load_workbook('OregonOriginal.xlsx')
    districtSheet = wb['District List']
    prevDistrict = ""
    inputRow = 0
    districtIndex = -1
    df = pd.DataFrame(
        columns=['district', 'on-site school count', 'hybrid school count',
                 'distance school count', 'distance w/LIPI school count', 'report date', 'date scraped'])

    for row in districtSheet.iter_rows(values_only=True):
        if inputRow == 0:  # Skip column headers
            inputRow += 1
            continue
        curDistrict = row[2]
        curMode = row[5]
        if curDistrict is None:  # End of input file
            break

        onSite = 0
        hybrid = 0
        distance = 0
        distanceLIPI = 0
        if curMode == "On-Site":
            onSite = 1
        elif curMode == "Hybrid":
            hybrid = 1
        elif curMode == "Comprehensive Distance Learning":
            distance = 1
        elif curMode == "Comprehensive Distance Learning w/LIPI":
            distanceLIPI = 1
        if curDistrict == prevDistrict:  # Continue with previous district
            if onSite == 1:
                df.iat[districtIndex, 1] = df.iat[districtIndex, 1] + 1
            elif hybrid == 1:
                df.iat[districtIndex, 2] = df.iat[districtIndex, 2] + 1
            elif distance == 1:
                df.iat[districtIndex, 3] = df.iat[districtIndex, 3] + 1
            elif distanceLIPI == 1:
                df.iat[districtIndex, 4] = df.iat[districtIndex, 4] + 1
        else:                              # Start new district
            districtIndex += 1
            reportWeek = row[3]
            dateUpdated = reportWeek[11:]  # End date of report week

            newDistrictRow = pd.Series(data={'district': curDistrict, 'on-site school count': onSite,
                                             'hybrid school count': hybrid, 'distance school count': distance,
                                             'distance w/LIPI school count': distanceLIPI, 'report date': dateUpdated,
                                             'date scraped': date.today()})

            df = df.append(newDistrictRow, ignore_index=True)

        inputRow += 1                      # End for
        prevDistrict = curDistrict
    df.to_csv('Oregon.csv', index=False)   # Copy dataframe to CSV

main()
